#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export Averaged Environmental Time Series per Event Configuration
=================================================================

Exports a multi-sheet Excel workbook with 1-minute averaged environmental
and particle time series for each unique test configuration. Each sheet
contains the mean, std, max, and min across all replicate events sharing the
same configuration, from shower-on through the 2-hour particle deposition window.

Flow rate grouping:
  - Standard flow (no FlowRate tag) + 4.1-5.6 LPM tagged events are lumped
    together under the base config key.
  - 1.4 LPM and 2.2 LPM events are reported as separate groups.

Sensor spaces:
  - Bath:     Vaisala MBa, HOBO Bathroom1, HOBO Bathroom2, Aranet4 Bathroom
  - Bath/Bed: HOBO Bath/Bed (separate column; transitional space)
  - Bedroom:  Vaisala Bed1, HOBO Bedroom1-3, Aranet4 Bedroom

Within each space the available sensors are averaged together to give one
value per event per minute, then statistics (mean/std/max/min) are computed
across replicate events.

Output:
  <data_root>/output/event_config_timeseries.xlsx

Usage:
  python scripts/export_config_timeseries.py [--no-sig-figs]

Author: Nathan Lima
Institution: National Institute of Standards and Technology (NIST)
Created: 2026-05-18
"""

import argparse
import re
import sys
import warnings
from datetime import timedelta
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

sys.path.insert(0, str(Path(__file__).parent.parent))

import src.sig_figs as sf  # noqa: E402
from scripts.event_registry import load_event_registry  # noqa: E402
from src.data_paths import get_data_root  # noqa: E402
from src.env_data_loader import SENSOR_CONFIG, load_quantaq_data, load_sensor_data  # noqa: E402
from src.event_manager import sort_config_keys_by_water_temp  # noqa: E402
from src.particle_calculations import PARTICLE_BINS  # noqa: E402

# =============================================================================
# Sensor Groups
# =============================================================================

BATH_RH_SENSORS = [
    "Vaisala MBa RH",
    "HOBO Bathroom1 RH",
    "HOBO Bathroom2 RH",
    "Aranet4 Bathroom RH",
]
BATH_TEMP_SENSORS = [
    "Vaisala MBa Temp",
    "HOBO Bathroom1 Temp",
    "HOBO Bathroom2 Temp",
    "Aranet4 Bathroom Temp",
]
BATH_BED_RH_SENSORS = ["HOBO Bath/Bed RH"]
BATH_BED_TEMP_SENSORS = ["HOBO Bath/Bed Temp"]
BEDROOM_RH_SENSORS = [
    "Vaisala Bed1 RH",
    "HOBO Bedroom1 RH",
    "HOBO Bedroom2 RH",
    "HOBO Bedroom3 RH",
    "Aranet4 Bedroom RH",
]
BEDROOM_TEMP_SENSORS = [
    "Vaisala Bed1 Temp",
    "HOBO Bedroom1 Temp",
    "HOBO Bedroom2 Temp",
    "HOBO Bedroom3 Temp",
    "Aranet4 Bedroom Temp",
]

# Ordered mapping: internal column name -> sensor list
SENSOR_GROUPS = {
    "bath_rh":       BATH_RH_SENSORS,
    "bath_temp":     BATH_TEMP_SENSORS,
    "bath_bed_rh":   BATH_BED_RH_SENSORS,
    "bath_bed_temp": BATH_BED_TEMP_SENSORS,
    "bedroom_rh":    BEDROOM_RH_SENSORS,
    "bedroom_temp":  BEDROOM_TEMP_SENSORS,
}

# All environmental sensors needed (for preloading)
ALL_ENV_SENSORS = (
    BATH_RH_SENSORS + BATH_TEMP_SENSORS
    + BATH_BED_RH_SENSORS + BATH_BED_TEMP_SENSORS
    + BEDROOM_RH_SENSORS + BEDROOM_TEMP_SENSORS
)

# Display label and unit for each environmental column
ENV_COLUMN_SPECS = [
    ("bath_rh",       "Bath RH",       "%"),
    ("bath_temp",     "Bath Temp",     "degC"),
    ("bath_bed_rh",   "Bath/Bed RH",   "%"),
    ("bath_bed_temp", "Bath/Bed Temp", "degC"),
    ("bedroom_rh",    "Bedroom RH",    "%"),
    ("bedroom_temp",  "Bedroom Temp",  "degC"),
]

# =============================================================================
# Flow Rate Grouping
# =============================================================================

FLOW_RATE_LUMP_MIN = 4.1
FLOW_RATE_LUMP_MAX = 5.6

_FLOW_RATE_RE = re.compile(r"_FlowRate([\d.]+)LPM")


def normalize_config_key(config_key: str) -> str:
    """Strip FlowRate tags in the 4.1-5.6 LPM range (treated as standard flow)."""
    m = _FLOW_RATE_RE.search(config_key)
    if m:
        rate = float(m.group(1))
        if FLOW_RATE_LUMP_MIN <= rate <= FLOW_RATE_LUMP_MAX:
            return config_key.replace(m.group(0), "")
    return config_key


# =============================================================================
# Excel Sheet Name Helpers
# =============================================================================

_INVALID_SHEET_CHARS = re.compile(r"[/\\?*\[\]:]")


def make_sheet_name(key: str, used: set) -> str:
    """Sanitize and truncate config key to a valid, unique Excel sheet name (<= 31 chars)."""
    sanitized = _INVALID_SHEET_CHARS.sub("_", key)
    base = sanitized[:31]
    if base not in used:
        return base
    for i in range(2, 100):
        candidate = sanitized[:28] + f"_{i:02d}"
        if candidate not in used:
            return candidate
    return sanitized[:31]


# =============================================================================
# Data Loading
# =============================================================================

def preload_sensor_data(events: list) -> dict:
    """
    Load all environmental sensor data for the full span of events.

    Args:
        events: List of event dicts with shower_on and deposition_end.

    Returns:
        Dict mapping sensor_name -> Series (DatetimeIndex, float values).
    """
    global_start = min(e["shower_on"] for e in events) - timedelta(minutes=5)
    global_end = max(e["deposition_end"] for e in events) + timedelta(minutes=5)

    print(f"  Date range: {global_start.date()} to {global_end.date()}")

    cache = {}
    for sensor_name in ALL_ENV_SENSORS:
        config = SENSOR_CONFIG.get(sensor_name)
        if config is None:
            print(f"    {sensor_name}: not in SENSOR_CONFIG, skipping")
            continue
        df = load_sensor_data(sensor_name, config, global_start, global_end)
        if df is not None and not df.empty:
            s = df.set_index("datetime")["value"].sort_index()
            cache[sensor_name] = s
            print(f"    {sensor_name}: {len(s):,} records")
        else:
            print(f"    {sensor_name}: no data found")

    return cache


def preload_particle_data(events: list) -> pd.DataFrame:
    """
    Load QuantAQ inside particle data for the full span of events.

    Args:
        events: List of event dicts with shower_on and deposition_end.

    Returns:
        DataFrame with DatetimeIndex and opc_bin0..opc_bin11 columns.
    """
    global_start = min(e["shower_on"] for e in events) - timedelta(minutes=5)
    global_end = max(e["deposition_end"] for e in events) + timedelta(minutes=5)

    df = load_quantaq_data("inside", global_start, global_end)
    if df.empty:
        print("  Warning: No inside particle data loaded.")
        return pd.DataFrame()

    df = df.set_index("datetime").sort_index()
    print(f"  Particle data: {len(df):,} records, {df.index.min()} to {df.index.max()}")
    return df


# =============================================================================
# Per-Event Time Series
# =============================================================================

def build_event_timeseries(
    event: dict,
    sensor_cache: dict,
    particle_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build a 1-minute time series for a single event indexed by minutes from shower_on.

    Sensors within the same space (bath, bath/bed, bedroom) are averaged together
    before returning so each space has one value per minute.

    Args:
        event: Event dict with shower_on and deposition_end.
        sensor_cache: Dict of sensor_name -> Series(DatetimeIndex, float).
        particle_df: DataFrame(DatetimeIndex) with opc_bin columns.

    Returns:
        DataFrame with integer minute index (0 = shower_on) and columns:
        bath_rh, bath_temp, bath_bed_rh, bath_bed_temp, bedroom_rh, bedroom_temp,
        bin0 .. bin11. Missing variables are absent (not NaN-padded).
    """
    shower_on = event["shower_on"]
    window_end = event["deposition_end"]

    columns = {}

    # --- Environmental sensors ---
    for group_name, sensor_list in SENSOR_GROUPS.items():
        sensor_series = []
        for sensor_name in sensor_list:
            if sensor_name not in sensor_cache:
                continue
            s = sensor_cache[sensor_name]
            s_win = s[(s.index >= shower_on) & (s.index <= window_end)]
            if s_win.empty:
                continue
            s_1min = s_win.resample("1min", origin=shower_on).mean()
            sensor_series.append(s_1min)

        if sensor_series:
            # Average across available sensors in this space
            space_avg = pd.concat(sensor_series, axis=1).mean(axis=1)
            columns[group_name] = space_avg

    # --- Particle bins ---
    if not particle_df.empty:
        p_win = particle_df[
            (particle_df.index >= shower_on) & (particle_df.index <= window_end)
        ]
        for bin_num, bin_info in PARTICLE_BINS.items():
            col = bin_info["column"]
            if col in p_win.columns:
                s = p_win[col].dropna()
                if not s.empty:
                    s_1min = s.resample("1min", origin=shower_on).mean()
                    columns[f"bin{bin_num}"] = s_1min

    if not columns:
        return pd.DataFrame()

    df = pd.DataFrame(columns)

    # Convert DatetimeIndex to integer minutes from shower_on
    df.index = ((df.index - shower_on).total_seconds() / 60).round(0).astype(int)
    df.index.name = "minute"

    # Drop any negative-minute rows (shouldn't occur given window slicing above)
    return df[df.index >= 0]


# =============================================================================
# Group Aggregation
# =============================================================================

def aggregate_group_events(event_dfs: list) -> pd.DataFrame:
    """
    Aggregate per-event DataFrames into mean / std / max / min per minute.

    Args:
        event_dfs: List of DataFrames, each indexed by minute offset.

    Returns:
        DataFrame with MultiIndex columns (variable, stat) indexed by minute.
    """
    if not event_dfs:
        return pd.DataFrame()

    combined = pd.concat(
        event_dfs, keys=range(len(event_dfs)), names=["event_idx", "minute"]
    )
    return combined.groupby("minute").agg(["mean", "std", "max", "min"])


def flatten_columns(agg: pd.DataFrame) -> pd.DataFrame:
    """
    Flatten MultiIndex columns to labelled strings with units.

    (bath_rh, mean) -> "Bath RH avg (%)"
    (bin3, std)     -> "Bin3 [1.0-1.3 um] std (#/cm3)"

    Args:
        agg: DataFrame from aggregate_group_events (MultiIndex columns).

    Returns:
        DataFrame with flat string columns and index renamed to
        "Minutes from shower on".
    """
    col_display = {col: display for col, display, _ in ENV_COLUMN_SPECS}
    col_unit    = {col: unit    for col, _, unit    in ENV_COLUMN_SPECS}

    for bin_num, bin_info in PARTICLE_BINS.items():
        key = f"bin{bin_num}"
        col_display[key] = f"Bin{bin_num} [{bin_info['name']} um]"
        col_unit[key]    = "#/cm3"

    stat_labels = {"mean": "avg", "std": "std", "max": "max", "min": "min"}

    new_cols = []
    for var, stat in agg.columns:
        display    = col_display.get(var, var)
        unit       = col_unit.get(var, "")
        stat_label = stat_labels.get(stat, stat)
        new_cols.append(
            f"{display} {stat_label} ({unit})" if unit else f"{display} {stat_label}"
        )

    result = agg.copy()
    result.columns = new_cols
    result.index.name = "Minutes from shower on"
    return result


# =============================================================================
# Excel Output
# =============================================================================

def write_excel(
    output_path: Path,
    groups: dict,
    group_order: list,
) -> None:
    """
    Write the multi-sheet Excel workbook.

    Sheet 1  : Index listing every configuration group with a hyperlink.
    Sheet 2+ : One sheet per group with 1-min averaged time series.

    Args:
        output_path: Destination .xlsx path.
        groups: Dict of norm_key -> (display_key, n_events, event_nums, flat_df).
        group_order: List of norm_keys in desired sheet order.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required. Install with: conda install -c conda-forge openpyxl"
        )

    wb = openpyxl.Workbook()
    header_font = Font(bold=True)
    link_font   = Font(color="0563C1", underline="single")

    # --- Assign sheet names up front (needed for hyperlinks) ---
    used_names: set = {"Index"}
    sheet_names: dict = {}
    for key in group_order:
        if key in groups:
            sheet_names[key] = make_sheet_name(key, used_names)
            used_names.add(sheet_names[key])

    # --- Index sheet ---
    ws_idx = wb.active
    ws_idx.title = "Index"

    idx_headers = ["#", "Config Key", "Sheet", "N Events", "Event Numbers"]
    ws_idx.append(idx_headers)
    for cell in ws_idx[1]:
        cell.font = header_font

    for i, key in enumerate((k for k in group_order if k in groups), start=1):
        display_key, n_events, event_nums, _ = groups[key]
        sname = sheet_names[key]
        event_nums_str = ", ".join(str(n) for n in sorted(event_nums))
        ws_idx.append([i, display_key, sname, n_events, event_nums_str])

        # Hyperlink on sheet-name cell
        link_cell = ws_idx.cell(row=ws_idx.max_row, column=3)
        link_cell.hyperlink = f"#'{sname}'!A1"
        link_cell.font = link_font

    ws_idx.column_dimensions["A"].width = 5
    ws_idx.column_dimensions["B"].width = 62
    ws_idx.column_dimensions["C"].width = 34
    ws_idx.column_dimensions["D"].width = 10
    ws_idx.column_dimensions["E"].width = 42

    # --- Data sheets ---
    for key in (k for k in group_order if k in groups):
        display_key, n_events, event_nums, flat_df = groups[key]
        sname = sheet_names[key]
        ws = wb.create_sheet(title=sname)

        # Metadata header rows
        ws.append([f"Config Key: {display_key}"])
        ws.cell(row=1, column=1).font = header_font
        event_nums_str = ", ".join(str(n) for n in sorted(event_nums))
        ws.append([f"N events: {n_events}    Event numbers: {event_nums_str}"])
        ws.append([])  # blank spacer row

        # Column header row
        col_headers = ["Minutes from shower on"] + list(flat_df.columns)
        ws.append(col_headers)
        for cell in ws[4]:
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Data rows
        for minute, row in flat_df.iterrows():
            ws.append([minute] + list(row))

        # Column widths
        ws.column_dimensions["A"].width = 24
        for col_idx in range(2, len(col_headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Freeze panes: keep header rows and minute column visible while scrolling
        ws.freeze_panes = ws.cell(row=5, column=2)

        # Row height for header row to show wrapped text
        ws.row_dimensions[4].height = 40

    wb.save(output_path)


# =============================================================================
# Main
# =============================================================================

def run(sig_figs_enabled: bool = True) -> None:
    """Load events, build per-group averages, and write the Excel workbook."""
    sf.set_enabled(sig_figs_enabled)

    data_root  = get_data_root()
    output_dir = data_root / "output"
    output_path = output_dir / "event_config_timeseries.xlsx"

    # --- Load event registry ---
    print("Loading event registry...")
    registry = load_event_registry()

    # Keep only non-excluded events with a valid event number and deposition window
    mask = (
        registry["event_number"].notna()
        & registry["deposition_end"].notna()
        & registry["shower_on"].notna()
    )
    if "is_excluded" in registry.columns:
        mask &= ~registry["is_excluded"].fillna(False)

    valid = registry[mask].copy()
    print(f"  {len(valid)} valid events found")

    if valid.empty:
        print("No valid events to process. Exiting.")
        return

    # Fill deposition_end from shower_off if missing (fallback)
    if valid["deposition_end"].isna().any():
        missing = valid["deposition_end"].isna()
        valid.loc[missing, "deposition_end"] = (
            valid.loc[missing, "shower_off"] + timedelta(hours=2)
        )

    # Normalize config keys for flow-rate grouping
    valid["_group_key"] = valid["config_key"].apply(normalize_config_key)

    events = valid.to_dict("records")

    # --- Preload data ---
    print("\nPreloading environmental sensor data...")
    sensor_cache = preload_sensor_data(events)

    print("\nPreloading particle data...")
    particle_df = preload_particle_data(events)

    # --- Build per-group output ---
    grouped = valid.groupby("_group_key")
    group_keys = list(grouped.groups.keys())

    # Sort by water temperature (primary) then alphabetically
    try:
        group_keys = sort_config_keys_by_water_temp(group_keys)
    except Exception:
        group_keys = sorted(group_keys)

    print(f"\nProcessing {len(group_keys)} configuration group(s)...")

    groups: dict = {}

    for norm_key in group_keys:
        group_rows = grouped.get_group(norm_key).to_dict("records")
        n_events   = len(group_rows)
        event_nums = [int(e["event_number"]) for e in group_rows]

        print(
            f"\n  [{norm_key}]  n={n_events}  events: {sorted(event_nums)}"
        )

        event_dfs = []
        for event in group_rows:
            df = build_event_timeseries(event, sensor_cache, particle_df)
            if df.empty:
                print(f"    Warning: no data for event {event['event_number']}")
            else:
                event_dfs.append(df)

        if not event_dfs:
            print("    No data available — group skipped.")
            continue

        agg     = aggregate_group_events(event_dfs)
        flat_df = flatten_columns(agg)

        if sig_figs_enabled:
            flat_df = sf.apply_sig_figs_to_df(flat_df)

        groups[norm_key] = (norm_key, n_events, event_nums, flat_df)

    if not groups:
        print("No groups produced output. Exiting.")
        return

    # --- Write Excel ---
    print(f"\nWriting {len(groups)} sheet(s) to {output_path} ...")
    write_excel(output_path, groups, group_keys)
    print(f"Done. Output: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Export averaged 1-minute environmental and particle time series "
            "per event configuration to Excel."
        )
    )
    parser.add_argument(
        "--no-sig-figs",
        action="store_true",
        help="Disable significant-figure rounding on output values.",
    )
    args = parser.parse_args()
    run(sig_figs_enabled=not args.no_sig_figs)


if __name__ == "__main__":
    main()
