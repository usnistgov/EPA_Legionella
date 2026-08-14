#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export Per-Sensor Event Configuration Time Series (MODULAIR-PM Fleet)
=====================================================================

Per-sensor variant of scripts/export_config_timeseries.py. Instead of the
single QuantAQ inside sensor, this processes each IAQ&V MODULAIR-PM fleet
sensor's raw chunk data (moduair_pm_chunks share) and, for every sensor,
generates:

  1. An aggregated workbook (same structure as the original single-sensor file):
     Index sheet plus one sheet per configuration group with the mean, std,
     max, and min per minute across replicate events, from 15 minutes before
     shower-on through the 2-hour deposition window (minute index 0 is
     shower-on, so the lead-in minutes are -15 through -1). Particle bins come
     from that sensor (1-minute grid, no smoothing); the bath/bath-bed/bedroom
     temperature and relative-humidity columns are shared across all sensor
     files.

  2. One raw workbook per configuration group. Each workbook has an Index sheet
     plus one sheet per replicate event. Every event sheet holds that sensor's
     raw particle records at native cadence (no 1-minute resample, no rolling
     average), from 15 minutes before shower-on through the 2-hour deposition
     window. The bath, bath/bed, and bedroom temperature and RH columns are
     appended after the bins, matched to each particle timestamp with a
     backward as-of merge (most recent reading within 2 minutes; blank
     otherwise).

Sensor set:
  All fleet sensors in the chunk share except MOD-PM-00467 and MOD-PM-00785:
  195, 401, 402, 465, 515, 516, 554, 555, 813, 814, 815, 816, 942, 943.

Event window:
  Only events whose shower-on time falls in 2026-06-03 through 2026-07-16 are
  processed. Sensors added to the test area late are gated to their install
  time (MOD-PM-00555: 2026-06-26 01:00; MOD-PM-00465, 00515, 00516:
  2026-07-08 13:00); events before a sensor's install time are skipped for
  that sensor only.

Flow rate grouping:
  Standard flow (no FlowRate tag) plus 4.1-5.6 LPM tagged events are lumped
  together under the base config key. 1.4 LPM and 2.2 LPM events are reported
  as separate groups (same as the original script).

Output:
  <data_root>/output/event_config_timeseries_fleet/
    MOD-PM-<sn>/MOD-PM-<sn>_event_config_timeseries.xlsx
    MOD-PM-<sn>/raw/<config_group>_raw.xlsx

Usage:
  python scripts/export_config_timeseries_fleet.py [--no-sig-figs] [--sensors 195 401 ...]

Author: Nathan Lima
Institution: National Institute of Standards and Technology (NIST)
Created: 2026-07-23
"""

import argparse
import sys
import warnings
from datetime import timedelta
from pathlib import Path

import pandas as pd

warnings.filterwarnings("ignore")

sys.path.insert(0, str(Path(__file__).parent.parent))

import src.sig_figs as sf  # noqa: E402
from scripts.event_registry import load_event_registry  # noqa: E402
from scripts.export_config_timeseries import (  # noqa: E402
    ENV_COLUMN_SPECS,
    SENSOR_GROUPS,
    aggregate_group_events,
    flatten_columns,
    make_sheet_name,
    normalize_config_key,
    preload_sensor_data,
    write_excel,
)
from src.data_paths import get_data_root  # noqa: E402
from src.event_manager import sort_config_keys_by_water_temp  # noqa: E402
from src.moduair_loader import (  # noqa: E402
    BIN_COLUMNS,
    _normalize_sn,
    load_sensor_bins,
    load_sensor_bins_raw,
)
from src.particle_calculations import PARTICLE_BINS  # noqa: E402

# =============================================================================
# Sensor Set and Gating
# =============================================================================

# The 14 IAQ&V fleet sensors: all sensors in the chunk share except 00467 and
# 00785. 00195 is included and handled like any other sensor (from chunk data).
SENSOR_IDS = [
    "00195",
    "00401",
    "00402",
    "00465",
    "00515",
    "00516",
    "00554",
    "00555",
    "00813",
    "00814",
    "00815",
    "00816",
    "00942",
    "00943",
]

# Global processing window (inclusive) on shower-on time.
DATE_START = pd.Timestamp("2026-06-03 00:00:00")
DATE_END = pd.Timestamp("2026-07-16 23:59:59")

# Per-sensor install cutoffs: an event is skipped for that sensor if its
# shower-on time is before the cutoff (sensor not yet in the test area).
SENSOR_INSTALL_CUTOFFS = {
    "00555": pd.Timestamp("2026-06-26 01:00:00"),
    "00465": pd.Timestamp("2026-07-08 13:00:00"),
    "00515": pd.Timestamp("2026-07-08 13:00:00"),
    "00516": pd.Timestamp("2026-07-08 13:00:00"),
}

# Environmental temperature/RH columns attached to the raw sheets, in output
# order (after the particle bins). The keys are SENSOR_GROUPS keys; the display
# label and unit reuse ENV_COLUMN_SPECS so the raw headers match the aggregated
# workbook. Raw env values are matched to each particle timestamp with a
# backward as-of merge (most recent reading), limited by ENV_ASOF_TOLERANCE.
ENV_RAW_ORDER = [
    "bath_temp",
    "bath_rh",
    "bath_bed_temp",
    "bath_bed_rh",
    "bedroom_temp",
    "bedroom_rh",
]
ENV_ASOF_TOLERANCE = pd.Timedelta(minutes=2)

# Pre-shower lead-in: exports begin this long before shower-on so the baseline
# just before the shower is captured. The aggregated minute index keeps 0 at
# shower-on, so these lead-in minutes are indexed -PRE_SHOWER_LEAD through -1.
PRE_SHOWER_LEAD = pd.Timedelta(minutes=15)


# =============================================================================
# Event Loading and Gating
# =============================================================================


def load_valid_events() -> pd.DataFrame:
    """
    Load the event registry and keep numbered events in the date window.

    Keeps events that have an event number, a deposition_end, and a shower_on
    time, then applies flow-rate grouping and restricts to the global date
    window on shower-on time. Unlike the single-sensor script, the registry
    ``is_excluded`` flag is not applied: it encodes CO2/RH data-quality criteria
    (lambda R^2, bedroom RH-mixing) that do not bear on PM records. Water-temp-
    testing runs are still dropped, since they never receive an event number.

    Returns:
        DataFrame of numbered events with a ``_group_key`` column, or empty.
    """
    print("Loading event registry...")
    registry = load_event_registry()

    mask = (
        registry["event_number"].notna()
        & registry["deposition_end"].notna()
        & registry["shower_on"].notna()
    )

    valid = registry[mask].copy()

    # Fill deposition_end from shower_off if missing (fallback, same as original)
    if valid["deposition_end"].isna().any():
        missing = valid["deposition_end"].isna()
        valid.loc[missing, "deposition_end"] = valid.loc[missing, "shower_off"] + timedelta(hours=2)

    # Restrict to the global date window on shower-on time
    in_window = (valid["shower_on"] >= DATE_START) & (valid["shower_on"] <= DATE_END)
    valid = valid[in_window].copy()

    # Normalize config keys for flow-rate grouping
    valid["_group_key"] = valid["config_key"].apply(normalize_config_key)

    print(f"  {len(valid)} valid events in {DATE_START.date()} to {DATE_END.date()}")
    return valid


def filter_events_for_sensor(valid: pd.DataFrame, sn: str) -> pd.DataFrame:
    """
    Apply the per-sensor install cutoff to the window-filtered events.

    Parameters:
        valid: Window-filtered events from load_valid_events().
        sn: Normalized 5-digit sensor ID.

    Returns:
        Subset of ``valid`` whose shower-on time is at or after the sensor's
        install cutoff (unchanged if the sensor has no cutoff).
    """
    cutoff = SENSOR_INSTALL_CUTOFFS.get(sn)
    if cutoff is None:
        return valid
    return valid[valid["shower_on"] >= cutoff].copy()


# =============================================================================
# Per-Event Time Series (this sensor's PM plus shared env sensors)
# =============================================================================


def build_event_pm_1min(
    event: dict,
    sensor_1min: pd.DataFrame,
    sensor_cache: dict,
) -> pd.DataFrame:
    """
    Build a 1-minute time series for one event: env sensors plus this sensor's bins.

    Environmental sensors within the same space are averaged together (same as
    the original build_event_timeseries). Particle bins come from this sensor's
    1-minute (no-smoothing) frame rather than the QuantAQ inside sensor.

    Parameters:
        event: Event dict with shower_on and deposition_end.
        sensor_1min: This sensor's DataFrame (datetime + opc_bin0..11) at 1-min.
        sensor_cache: Dict of env sensor_name -> Series(DatetimeIndex, float).

    Returns:
        DataFrame with integer minute index (0 = shower_on, negative values for
        the pre-shower lead-in) and env columns plus bin0..bin11. Missing
        variables are absent (not NaN-padded). Empty if the event has no data.
    """
    shower_on = event["shower_on"]
    window_start = shower_on - PRE_SHOWER_LEAD
    window_end = event["deposition_end"]

    columns = {}

    # --- Environmental sensors (shared groups, averaged per space) ---
    for group_name, sensor_list in SENSOR_GROUPS.items():
        sensor_series = []
        for sensor_name in sensor_list:
            if sensor_name not in sensor_cache:
                continue
            s = sensor_cache[sensor_name]
            s_win = s[(s.index >= window_start) & (s.index <= window_end)]
            if s_win.empty:
                continue
            s_1min = s_win.resample("1min", origin=shower_on).mean()
            sensor_series.append(s_1min)

        if sensor_series:
            space_avg = pd.concat(sensor_series, axis=1).mean(axis=1)
            columns[group_name] = space_avg

    # --- Particle bins from this sensor's 1-min frame ---
    if sensor_1min is not None and not sensor_1min.empty:
        p = sensor_1min.set_index("datetime").sort_index()
        p_win = p[(p.index >= window_start) & (p.index <= window_end)]
        for bin_num in PARTICLE_BINS:
            col = f"opc_bin{bin_num}"
            if col in p_win.columns:
                s = p_win[col].dropna()
                if not s.empty:
                    s_1min = s.resample("1min", origin=shower_on).mean()
                    columns[f"bin{bin_num}"] = s_1min

    if not columns:
        return pd.DataFrame()

    df = pd.DataFrame(columns)
    df.index = ((df.index - shower_on).total_seconds() / 60).round(0).astype(int)
    df.index.name = "minute"
    lead_min = int(PRE_SHOWER_LEAD.total_seconds() / 60)
    return df[df.index >= -lead_min]


def build_event_raw_pm(event: dict, sensor_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Slice this sensor's raw (native-cadence) particle records to the event window.

    Parameters:
        event: Event dict with shower_on and deposition_end.
        sensor_raw: This sensor's raw DataFrame (datetime + opc_bin0..11).

    Returns:
        DataFrame with a 'datetime' column and opc_bin0..11 columns for the
        window [shower_on - PRE_SHOWER_LEAD, deposition_end], sorted by time.
        Empty if no records.
    """
    if sensor_raw is None or sensor_raw.empty:
        return pd.DataFrame(columns=["datetime"] + BIN_COLUMNS)

    shower_on = event["shower_on"]
    window_start = shower_on - PRE_SHOWER_LEAD
    window_end = event["deposition_end"]
    mask = (sensor_raw["datetime"] >= window_start) & (sensor_raw["datetime"] <= window_end)
    return sensor_raw.loc[mask].sort_values("datetime").reset_index(drop=True)


def attach_env_columns(
    raw_df: pd.DataFrame,
    event: dict,
    sensor_cache: dict,
) -> pd.DataFrame:
    """
    Attach per-space temp/RH columns to a raw particle frame via as-of merge.

    For each environmental space (bath, bath/bed, bedroom temp and RH), the
    sensors in that group are averaged together on their native cadence over the
    event window (same space-averaging as build_event_pm_1min). Each averaged
    series is then matched to every raw particle timestamp with a backward
    merge_asof (most recent reading at or before the timestamp), limited to
    ENV_ASOF_TOLERANCE; matches older than the tolerance are left as NaN.

    Particle cadence is preserved: no env-only rows are added and the row count
    is unchanged. Columns are appended after the bins in ENV_RAW_ORDER; a group
    with no data in the window is omitted entirely.

    Parameters:
        raw_df: Frame with a 'datetime' column plus opc_bin0..11 (from
            build_event_raw_pm), sorted by time.
        event: Event dict with shower_on and deposition_end.
        sensor_cache: Dict of env sensor_name -> Series(DatetimeIndex, float).

    Returns:
        A copy of raw_df with the available env columns appended. Returned
        unchanged if raw_df is empty.
    """
    if raw_df is None or raw_df.empty:
        return raw_df

    shower_on = event["shower_on"]
    window_start = shower_on - PRE_SHOWER_LEAD
    window_end = event["deposition_end"]

    out = raw_df.sort_values("datetime").reset_index(drop=True).copy()
    times = out[["datetime"]]

    for group_name in ENV_RAW_ORDER:
        sensor_list = SENSOR_GROUPS.get(group_name, [])
        sensor_series = []
        for sensor_name in sensor_list:
            if sensor_name not in sensor_cache:
                continue
            s = sensor_cache[sensor_name]
            s_win = s[(s.index >= window_start) & (s.index <= window_end)]
            if s_win.empty:
                continue
            sensor_series.append(s_win)

        if not sensor_series:
            continue

        # Average replicate sensors in the space onto a common time index
        space_avg = pd.concat(sensor_series, axis=1).mean(axis=1).dropna().sort_index()
        if space_avg.empty:
            continue

        env_frame = pd.DataFrame(
            {"datetime": space_avg.index, group_name: space_avg.values}
        )
        merged = pd.merge_asof(
            times,
            env_frame,
            on="datetime",
            direction="backward",
            tolerance=ENV_ASOF_TOLERANCE,
        )
        out[group_name] = merged[group_name].values

    return out



# =============================================================================
# Raw Workbook Output (one per config group)
# =============================================================================


def write_raw_group_workbook(
    output_path: Path,
    group_key: str,
    event_sheets: list,
) -> None:
    """
    Write one raw workbook for a single configuration group.

    Sheet 1  : Index listing every event with its time span and record count.
    Sheet 2+ : One sheet per event with that sensor's raw particle records.

    Parameters:
        output_path: Destination .xlsx path.
        group_key: Configuration group key (for the header/title).
        event_sheets: List of (event_number, shower_on, deposition_end, raw_df),
            where raw_df has a 'datetime' column plus opc_bin0..11.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required. Install with: conda install -c conda-forge openpyxl"
        )

    wb = openpyxl.Workbook()
    header_font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")

    # Column labels for the bins (raw records: one column per analysis bin)
    bin_labels = {
        f"opc_bin{n}": f"Bin{n} [{PARTICLE_BINS[n]['name']} um] (#/cm3)" for n in PARTICLE_BINS
    }

    # Environmental temp/RH labels (bins-then-env order). Only the env columns
    # that appear in at least one event sheet are written, in ENV_RAW_ORDER.
    env_label = {key: f"{display} ({unit})" for key, display, unit in ENV_COLUMN_SPECS}
    present_env = [
        key
        for key in ENV_RAW_ORDER
        if any(key in raw_df.columns for _, _, _, raw_df in event_sheets)
    ]

    # --- Assign sheet names up front (event-number based) ---
    used_names: set = {"Index"}
    sheet_names: dict = {}
    for event_num, _, _, _ in event_sheets:
        base = make_sheet_name(f"Event_{event_num}", used_names)
        sheet_names[event_num] = base
        used_names.add(base)

    # --- Index sheet ---
    ws_idx = wb.active
    ws_idx.title = "Index"
    ws_idx.append([f"Config Group: {group_key}"])
    ws_idx.cell(row=1, column=1).font = header_font
    ws_idx.append([])
    ws_idx.append(["#", "Event Number", "Sheet", "Shower On", "Window End", "N Records"])
    for cell in ws_idx[3]:
        cell.font = header_font

    for i, (event_num, shower_on, window_end, raw_df) in enumerate(event_sheets, start=1):
        sname = sheet_names[event_num]
        ws_idx.append(
            [
                i,
                event_num,
                sname,
                shower_on.strftime("%Y-%m-%d %H:%M:%S"),
                window_end.strftime("%Y-%m-%d %H:%M:%S"),
                len(raw_df),
            ]
        )
        link_cell = ws_idx.cell(row=ws_idx.max_row, column=3)
        link_cell.hyperlink = f"#'{sname}'!A1"
        link_cell.font = link_font

    ws_idx.column_dimensions["A"].width = 5
    ws_idx.column_dimensions["B"].width = 14
    ws_idx.column_dimensions["C"].width = 20
    ws_idx.column_dimensions["D"].width = 22
    ws_idx.column_dimensions["E"].width = 22
    ws_idx.column_dimensions["F"].width = 12

    # --- Data sheets (one per event) ---
    for event_num, shower_on, window_end, raw_df in event_sheets:
        sname = sheet_names[event_num]
        ws = wb.create_sheet(title=sname)

        ws.append([f"Config Group: {group_key}    Event: {event_num}"])
        ws.cell(row=1, column=1).font = header_font
        ws.append([f"Window: {shower_on:%Y-%m-%d %H:%M:%S} to {window_end:%Y-%m-%d %H:%M:%S}"])
        ws.append([])

        col_headers = (
            ["Datetime"]
            + [bin_labels[c] for c in BIN_COLUMNS]
            + [env_label[k] for k in present_env]
        )
        ws.append(col_headers)
        for cell in ws[4]:
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for _, row in raw_df.iterrows():
            dt = row["datetime"]
            dt_str = dt.strftime("%Y-%m-%d %H:%M:%S") if pd.notna(dt) else ""
            bin_vals = [row[c] for c in BIN_COLUMNS]
            env_vals = [
                (row[k] if (k in raw_df.columns and pd.notna(row[k])) else None)
                for k in present_env
            ]
            ws.append([dt_str] + bin_vals + env_vals)

        ws.column_dimensions["A"].width = 22
        for col_idx in range(2, len(col_headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.freeze_panes = ws.cell(row=5, column=2)
        ws.row_dimensions[4].height = 40

    wb.save(output_path)


# =============================================================================
# Per-Sensor Processing
# =============================================================================


def process_sensor(
    sn: str,
    valid: pd.DataFrame,
    sensor_cache: dict,
    output_root: Path,
    sig_figs_enabled: bool,
) -> None:
    """
    Build and write the aggregated and raw workbooks for one sensor.

    Parameters:
        sn: Normalized 5-digit sensor ID.
        valid: Window-filtered events (before per-sensor cutoff).
        sensor_cache: Shared env sensor Series cache.
        output_root: event_config_timeseries_fleet/ directory.
        sig_figs_enabled: Apply sig-fig rounding to the aggregated workbook.
    """
    label = f"MOD-PM-{sn}"
    print(f"\n{'=' * 70}\n{label}\n{'=' * 70}")

    sensor_events = filter_events_for_sensor(valid, sn)
    if sensor_events.empty:
        print(f"  No events for {label} after install-cutoff gating; skipping.")
        return

    # Data span for this sensor's particle loads. The lower bound covers the
    # pre-shower lead-in plus a small margin so the earliest window is complete.
    span_start = sensor_events["shower_on"].min() - PRE_SHOWER_LEAD - timedelta(minutes=5)
    span_end = sensor_events["deposition_end"].max() + timedelta(minutes=5)

    print(f"  Loading 1-min bins ({span_start.date()} to {span_end.date()})...")
    sensor_1min = load_sensor_bins(sn, start=span_start, end=span_end, rolling_window_min=0)
    print(f"    1-min records: {len(sensor_1min):,}")

    print("  Loading raw (native-cadence) bins...")
    sensor_raw = load_sensor_bins_raw(sn, start=span_start, end=span_end)
    print(f"    raw records: {len(sensor_raw):,}")

    if sensor_1min.empty and sensor_raw.empty:
        print(f"  No particle data for {label}; skipping.")
        return

    # Group order sorted by water temperature (primary), then alphabetically
    group_keys = list(sensor_events["_group_key"].unique())
    try:
        group_keys = sort_config_keys_by_water_temp(group_keys)
    except Exception:
        group_keys = sorted(group_keys)

    # Output directories
    sensor_dir = output_root / label
    raw_dir = sensor_dir / "raw"
    sensor_dir.mkdir(parents=True, exist_ok=True)
    raw_dir.mkdir(parents=True, exist_ok=True)

    grouped = sensor_events.groupby("_group_key")
    agg_groups: dict = {}

    # Consistent raw-workbook filenames reuse the aggregated sheet names
    used_sheet_names: set = {"Index"}
    group_sheet_names: dict = {}
    for key in group_keys:
        group_sheet_names[key] = make_sheet_name(key, used_sheet_names)
        used_sheet_names.add(group_sheet_names[key])

    print(f"  Processing {len(group_keys)} configuration group(s)...")

    for norm_key in group_keys:
        group_rows = grouped.get_group(norm_key).to_dict("records")
        event_nums = [int(e["event_number"]) for e in group_rows]
        n_events = len(group_rows)

        event_dfs = []  # 1-min per-event frames for aggregation
        raw_event_sheets = []  # (event_num, shower_on, window_end, raw_df)

        for event in group_rows:
            pm_df = build_event_pm_1min(event, sensor_1min, sensor_cache)
            if not pm_df.empty:
                event_dfs.append(pm_df)

            raw_df = build_event_raw_pm(event, sensor_raw)
            raw_df = attach_env_columns(raw_df, event, sensor_cache)
            raw_event_sheets.append(
                (
                    int(event["event_number"]),
                    event["shower_on"],
                    event["deposition_end"],
                    raw_df,
                )
            )

        # --- Aggregated group ---
        if event_dfs:
            agg = aggregate_group_events(event_dfs)
            flat_df = flatten_columns(agg)
            if sig_figs_enabled:
                flat_df = sf.apply_sig_figs_to_df(flat_df)
            agg_groups[norm_key] = (norm_key, n_events, event_nums, flat_df)
        else:
            print(f"    [{norm_key}] no 1-min data; omitted from aggregated workbook.")

        # --- Raw workbook for this group ---
        if any(not r[3].empty for r in raw_event_sheets):
            raw_path = raw_dir / f"{group_sheet_names[norm_key]}_raw.xlsx"
            write_raw_group_workbook(raw_path, norm_key, raw_event_sheets)
            print(f"    [{norm_key}] raw workbook: {raw_path.name}")
        else:
            print(f"    [{norm_key}] no raw records; raw workbook skipped.")

    # --- Aggregated workbook ---
    if agg_groups:
        agg_path = sensor_dir / f"{label}_event_config_timeseries.xlsx"
        write_excel(agg_path, agg_groups, group_keys)
        print(f"  Aggregated workbook: {agg_path.name} ({len(agg_groups)} sheet(s))")
    else:
        print(f"  No aggregated groups produced for {label}.")


# =============================================================================
# Main
# =============================================================================


def run(sensor_ids: list, sig_figs_enabled: bool = True) -> None:
    """Load events and shared env data, then process each sensor."""
    sf.set_enabled(sig_figs_enabled)

    output_root = get_data_root() / "output" / "event_config_timeseries_fleet"

    valid = load_valid_events()
    if valid.empty:
        print("No valid events in window. Exiting.")
        return

    # Shared environmental sensor cache (loaded once for the full event span).
    # preload_sensor_data derives its window from shower_on with a 5-min margin;
    # shift the shower_on it sees back by the pre-shower lead-in so the cache
    # fully covers the earliest lead-in window.
    print("\nPreloading shared environmental sensor data...")
    cache_events = valid.copy()
    cache_events["shower_on"] = cache_events["shower_on"] - PRE_SHOWER_LEAD
    sensor_cache = preload_sensor_data(cache_events.to_dict("records"))

    for sid in sensor_ids:
        sn = _normalize_sn(sid)
        process_sensor(sn, valid, sensor_cache, output_root, sig_figs_enabled)

    print(f"\nDone. Output root: {output_root}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Export per-sensor averaged and raw event-configuration time series "
            "for the MODULAIR-PM fleet to Excel."
        )
    )
    parser.add_argument(
        "--no-sig-figs",
        action="store_true",
        help="Disable significant-figure rounding on aggregated values.",
    )
    parser.add_argument(
        "--sensors",
        nargs="+",
        default=None,
        help="Optional subset of sensor IDs to process (default: all 14).",
    )
    args = parser.parse_args()

    sensor_ids = args.sensors if args.sensors else SENSOR_IDS
    run(sensor_ids=sensor_ids, sig_figs_enabled=not args.no_sig_figs)


if __name__ == "__main__":
    main()
